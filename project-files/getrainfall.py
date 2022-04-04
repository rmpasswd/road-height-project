import openpyxl
from datetime import date

wb=openpyxl.load_workbook("./flood_data_new.xlsx")
shob=wb.get_sheet_by_name('Sheet1')


date = str(date.today())

asdf={}
for i in range(1, shob.max_row):

		if str(shob.cell(row=i,column=1).value) =date
			return shob.cell(row = i, column = 4).value

